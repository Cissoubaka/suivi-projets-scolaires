from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from odf.opendocument import OpenDocumentSpreadsheet
from odf.table import Table, TableRow, TableCell
from odf.text import P
import os

class ExcelExporter:
    def __init__(self, db):
        self.db = db

    def export_project(self, project_id, repetition=1, file_path=None):
        """Exporter les groupes d'un projet en Excel"""
        if file_path is None:
            project = self.db.get_project(project_id)
            file_path = f"{project[1]}_rep{repetition}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        project = self.db.get_project(project_id)
        groups = self.db.get_groups_for_project(project_id, repetition)

        wb = Workbook()
        ws = wb.active
        ws.title = "Groupes"

        # En-tête
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Titre du projet
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = f"Projet : {project[1]}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Informations du projet
        ws['A2'] = "Description :"
        ws['B2'] = project[2]
        ws['A3'] = "Répétition :"
        ws['B3'] = repetition
        ws['A4'] = "Taille des groupes :"
        ws['B4'] = project[4]

        # En-têtes du tableau
        row = 6
        ws[f'A{row}'].value = "Groupe"
        ws[f'B{row}'].value = "Élève 1"
        ws[f'C{row}'].value = "Élève 2"
        ws[f'D{row}'].value = "Élève 3"

        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{row}']
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Données des groupes
        row = 7
        for group in groups:
            group_id = group[0]
            group_num = group[2]
            students = self.db.get_students_in_group(group_id)

            ws[f'A{row}'].value = f"Groupe {group_num}"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = thin_border

            for idx, student in enumerate(students[:3]):  # Max 3 élèves
                col = chr(66 + idx)  # B, C, D
                ws[f'{col}{row}'].value = f"{student[2]} {student[1]}"  # lastname firstname
                ws[f'{col}{row}'].border = thin_border

            row += 1

        # Ajuster les largeurs des colonnes
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25

        wb.save(file_path)
        return file_path

    def export_all_projects(self, file_path=None):
        """Exporter tous les projets avec leurs groupes"""
        if file_path is None:
            file_path = f"ALL_PROJECTS_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        projects = self.db.get_all_projects()

        wb = Workbook()
        wb.remove(wb.active)  # Supprimer la feuille vide par défaut

        for project in projects:
            project_id = project[0]
            project_name = project[1]
            repetitions = project[3]

            for rep in range(1, repetitions + 1):
                ws = wb.create_sheet(f"{project_name}_{rep}")
                groups = self.db.get_groups_for_project(project_id, rep)

                # Titre
                ws.merge_cells('A1:D1')
                title_cell = ws['A1']
                title_cell.value = f"Projet : {project_name} (Rép. {rep})"
                title_cell.font = Font(bold=True, size=12)
                title_cell.alignment = Alignment(horizontal='center')

                # En-têtes
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                headers = ['Groupe', 'Élève 1', 'Élève 2', 'Élève 3']
                for col_idx, header in enumerate(headers):
                    cell = ws.cell(row=3, column=col_idx + 1)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = thin_border

                # Données
                row = 4
                for group in groups:
                    group_id = group[0]
                    group_num = group[2]
                    students = self.db.get_students_in_group(group_id)

                    ws.cell(row=row, column=1).value = f"Groupe {group_num}"
                    ws.cell(row=row, column=1).border = thin_border

                    for idx, student in enumerate(students[:3]):
                        ws.cell(row=row, column=idx + 2).value = f"{student[2]} {student[1]}"
                        ws.cell(row=row, column=idx + 2).border = thin_border

                    row += 1

                # Ajuster les largeurs
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 25
                ws.column_dimensions['C'].width = 25
                ws.column_dimensions['D'].width = 25

        wb.save(file_path)
        return file_path


class ODSExporter:
    """Exporte les groupes et données en fichiers ODS (LibreOffice Calc)"""
    
    def __init__(self, db):
        self.db = db
    
    def _add_simple_cell(self, row, text, valuetype="string", value_attr=None):
        """Ajoute une cellule simple à une ligne"""
        if valuetype != "string" and value_attr is not None:
            cell = TableCell(valuetype=valuetype, value=str(value_attr))
        else:
            cell = TableCell(valuetype=valuetype)
        
        p = P(text=str(text) if text else "")
        cell.addElement(p)
        row.addElement(cell)
        return cell

    def export_group_to_ods(self, group_id, prefix="Evaluation", dest_path=None):
        """Exporter un groupe vers un fichier ODS avec plusieurs onglets"""
        group = self.db.get_group_by_id(group_id)
        if not group:
            raise ValueError(f"Groupe {group_id} non trouvé")
        
        project_id = group[1]
        repetition = group[3]
        group_num = group[2]
        project = self.db.get_project(project_id)
        
        # Déterminer le chemin de destination
        if dest_path is None or dest_path == ".":
            dest_path = os.getcwd()
        else:
            # Convertir en chemin absolu si relatif
            dest_path = os.path.abspath(dest_path)
        
        # Créer le répertoire s'il n'existe pas
        if not os.path.exists(dest_path):
            os.makedirs(dest_path, exist_ok=True)
        
        # Créer le nom du fichier
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_name = f"{prefix}_Groupe{group_num}_{timestamp}.ods"
        file_path = os.path.join(dest_path, file_name)
        
        # Créer le document ODS
        doc = OpenDocumentSpreadsheet()
        
        # Onglet 1: Suivi de Présence
        self._create_attendance_sheet(doc, group_id, group_num, project, repetition)
        
        # Onglet 2: Évaluation
        self._create_evaluation_sheet(doc, group_id, group_num, project)
        
        # Sauvegarder le document
        doc.save(file_path)
        return file_path

    def _create_attendance_sheet(self, doc, group_id, group_num, project, repetition):
        """Créer l'onglet de suivi de présence"""
        table = Table(name="Suivi")
        doc.spreadsheet.addElement(table)
        
        # Titre
        row = TableRow()
        table.addElement(row)
        cell = TableCell(valuetype="string")
        cell.addElement(P(text=f"Suivi de Présence - Groupe {group_num}"))
        row.addElement(cell)
        
        # Info projet
        row = TableRow()
        table.addElement(row)
        cell = TableCell(valuetype="string")
        cell.addElement(P(text=f"Projet : {project[1]}"))
        row.addElement(cell)
        
        # Ligne vide
        row = TableRow()
        table.addElement(row)
        cell = TableCell(valuetype="string")
        cell.addElement(P(text=""))
        row.addElement(cell)
        
        # En-têtes principaux (Élève, Classe, puis dates)
        sessions = self.db.get_session_dates(project[0], repetition)
        
        # Première ligne d'en-têtes: Élève, Classe, puis dates
        row = TableRow()
        table.addElement(row)
        
        self._add_simple_cell(row, "Élève")
        self._add_simple_cell(row, "Classe")
        
        for session in sessions:
            # Date dans la première colonne du groupe de 4
            self._add_simple_cell(row, session[1])
            
            # 3 cellules vides pour les sous-colonnes
            for _ in range(3):
                self._add_simple_cell(row, "")
        
        # Ajouter une cellule pour le Total Suivi
        self._add_simple_cell(row, "")
        
        # Deuxième ligne pour les sous-colonnes d'évaluation
        row = TableRow()
        table.addElement(row)
        
        self._add_simple_cell(row, "")
        self._add_simple_cell(row, "")
        
        for session in sessions:
            for label in ["Prés.", "Journal", "Gantt", "Trav."]:
                self._add_simple_cell(row, label)
        
        # Ajouter le label pour le Total Suivi
        self._add_simple_cell(row, "Total Suivi")
        
        # Ajouter les élèves et leurs présences/notes
        students = self.db.get_students_in_group(group_id)
        
        for student in students:
            row = TableRow()
            table.addElement(row)
            
            # Nom élève
            student_name = f"{student[1]} {student[2]}"
            self._add_simple_cell(row, student_name)
            
            # Classe
            class_name = ""
            if student[3]:
                class_obj = self.db.get_class_by_id(student[3])
                if class_obj:
                    class_name = class_obj[1]
            self._add_simple_cell(row, class_name)
            
            # Calculer le total global pour tous les sessions
            total_global = 0
            
            # Présence et notes pour chaque séance
            for session in sessions:
                attendance = self.db.get_attendance(student[0], group_id, session[0])
                # attendance = (present, journal_bord, gantt, travail_comportement)
                present = attendance[0]
                journal_bord = attendance[1]
                gantt = attendance[2]
                travail_comportement = attendance[3]
                
                # Ajouter à la somme globale
                total_global += journal_bord + gantt + travail_comportement
                
                # Présence
                presence_str = "✓" if present else "✗"
                self._add_simple_cell(row, presence_str)
                
                # Journal de Bord
                self._add_simple_cell(row, str(journal_bord), valuetype="float", value_attr=journal_bord)
                
                # Gantt
                self._add_simple_cell(row, str(gantt), valuetype="float", value_attr=gantt)
                
                # Travail et Comportement
                self._add_simple_cell(row, str(travail_comportement), valuetype="float", value_attr=travail_comportement)
            
            # Total du suivi (somme de toutes les notes)
            self._add_simple_cell(row, str(total_global), valuetype="float", value_attr=total_global)

    def _create_evaluation_sheet(self, doc, group_id, group_num, project):
        """Créer l'onglet d'évaluation avec structure hiérarchique"""
        table = Table(name="Évaluation")
        doc.spreadsheet.addElement(table)
        
        # Récupérer la répétition du groupe
        group = self.db.get_group_by_id(group_id)
        repetition = group[3]
        
        # Titre
        row = TableRow()
        table.addElement(row)
        cell = TableCell(valuetype="string")
        cell.addElement(P(text=f"Évaluation - Groupe {group_num}"))
        row.addElement(cell)
        
        # Info projet
        row = TableRow()
        table.addElement(row)
        cell = TableCell(valuetype="string")
        cell.addElement(P(text=f"Projet : {project[1]}"))
        row.addElement(cell)
        
        # Ligne vide
        row = TableRow()
        table.addElement(row)
        cell = TableCell(valuetype="string")
        cell.addElement(P(text=""))
        row.addElement(cell)
        
        # Récupérer la hiérarchie des catégories
        categories_level1 = self.db.get_rating_categories(project[0])
        
        # Construire la liste des colonnes à afficher (Level 2 ou Level 3)
        # Structure: [(col_id, col_name, cat2_name, cat1_name), ...]
        col_list = []
        for cat1 in categories_level1:
            cat1_id, cat1_name, cat1_points = cat1
            categories_level2 = self.db.get_rating_subcategories(cat1_id)
            for cat2 in categories_level2:
                cat2_id, cat2_name, cat2_points = cat2
                categories_level3 = self.db.get_rating_subsubcategories(cat2_id)
                
                if categories_level3:
                    # S'il y a des Level 3, les ajouter tous
                    for cat3 in categories_level3:
                        cat3_id, cat3_name, cat3_points = cat3
                        col_list.append((cat3_id, cat3_name, cat2_name, cat1_name))
                else:
                    # S'il n'y a pas de Level 3, ajouter le Level 2 lui-même
                    col_list.append((cat2_id, cat2_name, cat2_name, cat1_name))
        
        # Première ligne d'en-têtes: Catégories niveau 1 et 2
        row = TableRow()
        table.addElement(row)
        
        self._add_simple_cell(row, "Élève")
        self._add_simple_cell(row, "Classe")
        self._add_simple_cell(row, "total_max")
        self._add_simple_cell(row, "TOTAL")
        self._add_simple_cell(row, "Suivi")
        
        for col_id, col_name, cat2_name, cat1_name in col_list:
            # Afficher la hiérarchie Cat1
            self._add_simple_cell(row, f"{cat1_name}")
        
        # Deuxième ligne d'en-têtes: Catégories niveau 3
        row = TableRow()
        table.addElement(row)
        
        # Cellules vides pour Élève, Classe, total_max, TOTAL, Suivi
        for _ in range(5):
            self._add_simple_cell(row, "")
        
        for col_id, col_name, cat2_name, cat1_name in col_list:
            # Afficher Cat2 / ColName (Col peut être Cat3 ou Cat2)
            if col_name == cat2_name:
                # C'est une categorie Level 2 sans Level 3
                label = f"{cat2_name}"
            else:
                # C'est une categorie Level 3
                label = f"{cat2_name} / {col_name}"
            
            self._add_simple_cell(row, label)
        
        # Ajouter les élèves et leurs évaluations
        students = self.db.get_students_in_group(group_id)
        for student in students:
            row = TableRow()
            table.addElement(row)
            
            # Nom élève
            student_name = f"{student[1]} {student[2]}"
            self._add_simple_cell(row, student_name)
            
            # Classe
            class_name = ""
            if student[3]:
                class_obj = self.db.get_class_by_id(student[3])
                if class_obj:
                    class_name = class_obj[1]
            self._add_simple_cell(row, class_name)
            
            # Calculer le maximum d'évaluation et de suivi
            # Maximum de suivi = somme des maximums de chaque session
            max_suivi = 0
            sessions = self.db.get_session_dates(project[0], repetition)
            for session_id, session_date in sessions:
                cursor = self.db.get_connection().cursor()
                cursor.execute('SELECT journal_bord, gantt, travail_comportement FROM session_dates WHERE id = ?', (session_id,))
                result = cursor.fetchone()
                cursor.connection.close()
                if result:
                    max_suivi += (result[0] or 0) + (result[1] or 0) + (result[2] or 0)
            
            # Maximum d'évaluation = somme des points des catégories assignées
            max_eval = 0
            for col_id, col_name, cat2_name, cat1_name in col_list:
                # Vérifier si la catégorie est assignée à l'élève
                is_assigned = self.db.get_student_rating_assignment(student[0], group_id, col_id)
                if is_assigned is None:
                    is_assigned = True
                
                if is_assigned:
                    # Récupérer les points de la catégorie
                    cursor = self.db.get_connection().cursor()
                    cursor.execute('SELECT points FROM rating_categories WHERE id = ?', (col_id,))
                    result = cursor.fetchone()
                    cursor.connection.close()
                    if result and result[0]:
                        max_eval += result[0]
            
            total_max = max_eval + max_suivi
            self._add_simple_cell(row, str(total_max), valuetype="float", value_attr=total_max)
            
            # SUIVI (somme de toutes les notes de présence: journal_bord + gantt + travail_comportement)
            suivi_score = 0
            sessions = self.db.get_session_dates(project[0], repetition)
            for session_id, session_date in sessions:
                attendance = self.db.get_attendance(student[0], group_id, session_id)
                journal_bord = attendance[1]
                gantt = attendance[2]
                travail_comportement = attendance[3]
                suivi_score += journal_bord + gantt + travail_comportement
            
            # TOTAL (somme de toutes les notes d'évaluation + suivi)
            total_score = 0
            for col_id, col_name, cat2_name, cat1_name in col_list:
                # Déterminer le niveau de la catégorie
                if col_name == cat2_name:
                    # C'est une catégorie Level 2
                    note = self.db.get_student_evaluation(student[0], group_id, col_id, rating_level=2)
                else:
                    # C'est une catégorie Level 3
                    note = self.db.get_student_evaluation(student[0], group_id, col_id, rating_level=3)
                total_score += note
            
            # Ajouter le suivi au total
            total_score += suivi_score
            
            self._add_simple_cell(row, str(total_score), valuetype="float", value_attr=total_score)
            self._add_simple_cell(row, str(suivi_score), valuetype="float", value_attr=suivi_score)
            
            # Notes pour chaque colonne
            for col_id, col_name, cat2_name, cat1_name in col_list:
                # Déterminer le niveau de la catégorie
                if col_name == cat2_name:
                    # C'est une catégorie Level 2
                    note = self.db.get_student_evaluation(student[0], group_id, col_id, rating_level=2)
                else:
                    # C'est une catégorie Level 3
                    note = self.db.get_student_evaluation(student[0], group_id, col_id, rating_level=3)
                self._add_simple_cell(row, str(note) if note > 0 else "", valuetype="float", value_attr=note)

